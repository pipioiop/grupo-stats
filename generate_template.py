import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "Setup"

# --- MATCH INFO (Rows 1-10) ---
# Column A must contain these exact keys for the parser to work
match_keys = [
    "Match Date", "Location", "Competition", "Stage", 
    "Kick Off Time", "Home Team", "Away Team", 
    "Home Display", "Away Display", "Team Analyzed"
]
# Column B: Sample Values (User edits these)
match_values = [
    "2024-02-01", "Pavilhão", "Liga", "Jornada 1", 
    "16:00", "Equipa Casa", "Equipa Fora", 
    "CASA", "FORA", "Equipa Casa"
]

for i, (key, val) in enumerate(zip(match_keys, match_values), start=1):
    ws.cell(row=i, column=1, value=key).font = Font(bold=True)
    ws.cell(row=i, column=2, value=val)

# --- PLAYERS (Rows 14-29) ---
# Header at Row 13 for reference
headers = ["Player ID", "Number", "First Name", "Last Name", "Position"]
for col, h in enumerate(headers, start=1):
    cell = ws.cell(row=13, column=col, value=h)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

# Sample Players (16 slots)
# The app reads 16 rows starting at 14.
players = [
    [1, 1, "João", "Silva", "GK"],
    [2, 2, "Antonio", "Costa", "FIXO"],
    [3, 3, "Manuel", "Pereira", "ALA"],
    [4, 4, "Jose", "Santos", "ALA"],
    [5, 5, "Luis", "Ferreira", "PIVOT"],
    [6, 6, "Carlos", "Oliveira", "PIVOT"],
    [7, 7, "Paulo", "Rodrigues", "ALA"],
    [8, 8, "Pedro", "Martins", "ALA"],
    [9, 12, "Miguel", "Gomes", "GK"],
    [10, 10, "Rui", "Fernandes", "ALA"],
    [11, 11, "Tiago", "Goncalves", "ALA"],
    [12, 13, "David", "Almeida", "FIXO"],
    [13, 14, "Ricardo", "Ribeiro", "PIVOT"],
    [14, 15, "Jorge", "Pinto", "ALA"],
    [15, 16, "Daniel", "Carvalho", "ALA"],
    [16, 17, "Andre", "Lopes", "FIXO"]
]

for i, p in enumerate(players, start=14):
    for col, val in enumerate(p, start=1):
        ws.cell(row=i, column=col, value=val)

# Save
filename = "Template_Jogo_Futsal.xlsx"
wb.save(filename)
print(f"Created {filename}")
