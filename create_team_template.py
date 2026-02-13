import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "Team Template"

# Match Info Data
match_info = [
    ["Match Date", "2024-02-01"],
    ["Location", "Pavilhão da Luz"],
    ["Competition", "Liga Placard"],
    ["Stage", "Fase Regular"],
    ["Kick Off Time", "16:00"],
    ["Home Team", "Benfica"],
    ["Away Team", "Sporting"],
    ["Home Display", "SLB"],
    ["Away Display", "SCP"],
    ["Team Analyzed", "Benfica"]
]

# Write Match Info to A1:B10
for i, row in enumerate(match_info, start=1):
    ws.cell(row=i, column=1, value=row[0])
    ws.cell(row=i, column=2, value=row[1])

# Player Info Headers (Optional, but good for context, code expects data at row 14)
headers = ["ID", "Number", "First Name", "Last Name", "Position"]
for col, header in enumerate(headers, start=1):
    ws.cell(row=13, column=col, value=header)

# Player Data (16 players)
players = [
    [1, 1, "André", "Sousa", "GK"],
    [2, 2, "Afonso", "Jesus", "FIXO"],
    [3, 3, "Bruno", "Coelho", "ALA"],
    [4, 8, "Gonçalo", "Sobral", "ALA"],
    [5, 10, "Arthur", "Guilherme", "ALA"],
    [6, 70, "Vinicius", "Rocha", "PIVOT"],
    [7, 9, "Jacaré", "Souza", "PIVOT"],
    [8, 11, "Chishkala", "Ivan", "ALA"],
    [9, 12, "Gugiel", "Léo", "GK"],
    [10, 13, "Diego", "Nunes", "ALA"],
    [11, 14, "Lúcio", "Rocha", "ALA"],
    [12, 18, "Carlos", "Monteiro", "ALA"],
    [13, 22, "Silvestre", "Ferreira", "ALA"],
    [14, 99, "Higor", "Souza", "PIVOT"],
    [15, 20, "Edmilson", "Kutchy", "ALA"],
    [16, 17, "Gustavo", "Marques", "FIXO"]
]

# Write Player Data to A14:E29
for i, player in enumerate(players, start=14):
    ws.cell(row=i, column=1, value=player[0]) # PID
    ws.cell(row=i, column=2, value=player[1]) # PNO
    ws.cell(row=i, column=3, value=player[2]) # First
    ws.cell(row=i, column=4, value=player[3]) # Last
    ws.cell(row=i, column=5, value=player[4]) # Pos

# Save the file
wb.save("exemplo_equipa.xlsx")
print("exemplo_equipa.xlsx created successfully.")
